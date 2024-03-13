import os 
import openpyxl
import sqlite3
import json
import random
from values.values import names, emails, telephone, cityByState, state
from time import sleep

class questionsToGenerateData:
    # Faz as perguntas e retorna o valor e confere se o valor é valido
    # No fim retorna um dicionario com os valores 
    def __init__(self):
        if not self.values_selected() or not self.file_type_selected() or not self.amount_selected():
            return  # Sai do método __init__ e interrompe a execução do programa

        self.return_data()
    def values_selected(self):
        values = str(input('''
Informe os valores que deseja gerar
1 - Nome
2 - Telefone   
3 - Email
4 - Estado
5 - Cidade
Exemplo: 1,3,4
'''))  
        # Converter a resposta para uma lista de string com os valores selecionados
        listValues = {
            '1': 'names',
            '2': 'telephone',   
            '3': 'emails',
            '4': 'state',
            '5': 'cityByState'
        }
        values_selectedStr = list(values.replace(',', ''))
        values_selectedStr_sorted = sorted(values_selectedStr, key=lambda x: int(x))
        # Verifica se o valor selecionado é válido 
        if all(value in listValues for value in values_selectedStr):
            self.filtredValue = {key: value for key, value in listValues.items() if key in values_selectedStr_sorted}
            return True
        else:
            return False
    def file_type_selected(self):
        self.fileType = int(input('''
Você deseja salvar os dados gerados onde?
1 - txt
2 - db(sqlite3)
3 - xlsx
4 - JSON
'''))
        # Converte a resposta para uma lista de valores, logo após verifica se o valor escolhido é valido 
        listFileType = {
            1: 'txt',
            2: 'db',   
            3: 'xlsx',
            4: 'JSON'
        }
    
        if self.fileType in listFileType:
            filtredValue = {key: value for key, value in listFileType.items() if key == self.fileType}
            return filtredValue
        else:
            return False
    def amount_selected(self):
        self.amount = int(input('Quantos dados devem ser gerados? obs: Quanto maior a quantidade mais tempo levará\n'))   
    def return_data(self):
        data = {
        'values': self.filtredValue,
        'fileType': self.fileType,
        'amount': self.amount
        }
        return data
    
class createTableSqlite3:
    # Responsavel pela logica de criação do banco de dado, tornando dinamico com base
    # nas opções escolhidas
    def __init__(self, values, amount):
        self.db_connect()
        self.create_table(values)
        self.fill_table(values, amount)
        pass
    def db_connect(self):
        #  cria o arquivo data.db e conecta ao banco
        self.db_name = str(input('Qual o nome do banco de dados?\n'))
        self.connection = sqlite3.connect('data.db')
        return self.connection   
    def create_table(self, values):
        # Cria as tabelas com base nos valores escolhidos anteriormente
        self.cursor = self.connection.cursor()
        # Usar um if dentro para preecher a tabela de forma dinamica
        columns = ['id INTEGER PRIMARY KEY']
        self.columnsString = ['id']
        if 'names' in values.values():
            columns.append('name TEXT NOT NULL')
            self.columnsString.append('name')
        if 'emails' in values.values():
            columns.append('email TEXT NOT NULL')
            self.columnsString.append('email')
        if 'telephone' in values.values():
            columns.append('telephone TEXT NOT NULL')
            self.columnsString.append('telephone')
        if 'state' in values.values():
            columns.append('state TEXT NOT NULL')
            self.columnsString.append('state')
        if 'cityByState' in values.values():
            columns.append('city TEXT NOT NULL')
            self.columnsString.append('city')

        self.cursor.execute(f'''
            CREATE TABLE {self.db_name} (
                {', '.join(columns)}
            );
        ''')
        self.connection.commit()   
    def fill_table(self, values, amount):
        # Preenche a tabela com valores aleatorios
        id = 0    
        for _ in range(amount):
            for value in values.values():
                id += 1
                if value == 'names':
                    name = random.choice(names)
                elif value == 'telephone':
                    self.telephone = random.choice(telephone)
                elif value == 'emails':
                    email = random.choice(emails)
                elif value == 'state':
                    self.state = random.choice(state)
                elif value == 'cityByState':
                    if 'state' in values.values():
                        city=(random.choice(cityByState[self.state]))
                    else:
                        city=(random.choice(cityByState[random.choice(state)]))
            # Pega todos os valores e armazena em uma lista, para então passar os valores
            # Como uma tupla ao banco de dados
            valores = [id]
            valores.append(name) if name else None
            valores.append(self.telephone) if self.telephone else None
            valores.append(email) if email else None
            valores.append(self.state) if self.state else None
            valores.append(city) if city else None
            print(len(valores))
            # Preenche com '?' para o número correto de placeholders
            placeholders = ', '.join(['?' for _ in range(len(valores))])
            self.cursor.execute(f'insert into {self.db_name} values({placeholders})', tuple(valores))
        # Encerra a operação
        self.connection.commit()
        self.connection.close()

def generate_data_json_txt(values, amount):
    # Pega o valor e a quantidade desejada, fazendo o tratamento dos dados e deixando em um formato que facilite
    # A visualização
    listData = []
    
    for _ in range(amount):
        dictData = {}
        for value in values.values():
            if value == 'names':
                dictData['name']=(random.choice(names))
            if value == 'telephone':
                dictData['telephone']=(random.choice(telephone))
            if value == 'emails':
                dictData['email']=(random.choice(emails))
            if value == 'state':
                dictData['state']=(random.choice(state))
            if value == 'cityByState':
                # Escolhe a cidade com base no estado, caso não tenha seleciona aleatoriamente
                if 'state' in values.values():
                    dictData['city']=(random.choice(cityByState[dictData['state']]))
                else:
                    dictData['city']=(random.choice(cityByState[random.choice(state)]))
        listData.append(dictData)          
    return listData 
def generate_data_txt(values, amount):
    data = generate_data_json_txt(values, amount)
    with open('data.txt', 'a', newline='', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
def generate_data_json(values, amount):
    data = generate_data_json_txt(values, amount)
    with open('data.json', 'a', newline='', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

def generate_data_db(values, amount):
    createTableSqlite3(values, amount)


def generate_random_data(values, amount):
    data = []
    print(values.values())
    for _ in range(amount):
        name = random.choice(names)
        email = random.choice(emails)
        phone = random.choice(telephone)
        state_choice = random.choice(state)
        city = random.choice(cityByState[state_choice])
        data.append((name if 'names' in values.values() else None , 
                     email if 'emails' in values.values() else None, 
                     phone if 'telephone' in values.values() else None, 
                     state_choice if 'state' in values.values() else None, 
                     city if 'cityByState' in values.values() else None))

    return data

def create_excel_file(values ,amount):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Random Data"

    headers = []
    if 'names' in values.values():
        headers.append('Name')
    if 'emails' in values.values():
        headers.append('Emails')
    if 'telephone' in values.values():
        headers.append('Telephone')
    if 'state' in values.values():
        headers.append('State')
    if 'cityByState' in values.values():
        headers.append('City')
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    random_data = generate_random_data(values, amount)
    for row, row_data in enumerate(random_data, start=2):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col, value=value)

    workbook.save('data.xlsx')

def generate_data_xlsx(values, amount):
    create_excel_file(values, amount)
    

def main():
    #  Print de apresentação do programa
    print('''
Olá, seja bem vindo ao gerador de dados para teste. Adiante serão
feitas algumas perguntas para que o gerador funcione de acordo com o seu objetivo.
Todas as perguntas tem um valor padrão definido, então caso seja enviado
uma resposta em branco o valor será preenchido para o padrão. No caso de perguntas que permitem
multipla escolha, separe cada valor com uma ","(virgula), sem a necessidade de espaço após a ","(virgula)''')
    
    #  Variables
    stop = False 
    # Variables from input

    while True:
        try:
            startingApp = str(input('Deseja inciar o gerador de dados falso?  não(encerra o programa)/sim \n'))
            os.system('cls')
            if startingApp == 'não' or startingApp == "nao":
                return False
            elif startingApp == 'sim':
                #  Pegar os valores para gerar o conteúdo        
                questions = questionsToGenerateData().return_data()
                # Gerar os dados no arquivo escolhido
                if questions['fileType'] == 1:
                    generate_data_txt(questions['values'], questions['amount'])
                elif questions['fileType'] == 2:
                    generate_data_db(questions['values'], questions['amount'])
                elif questions['fileType'] == 3:
                    generate_data_xlsx(questions['values'], questions['amount'])
                elif questions['fileType'] == 4:
                    generate_data_json(questions['values'], questions['amount'])
            else:
                print('Digite um valor válido')
        except:
            print('Digite um valor válido')  

        
main()